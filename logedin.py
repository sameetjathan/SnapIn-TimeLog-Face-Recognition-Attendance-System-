import io
import tkinter as tk
from tkinter import StringVar, ttk
import customtkinter as ctk
import os
import mysql.connector
from PIL import Image, ImageTk, ImageFilter, ImageDraw
from css import *
from datetime import datetime
import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from FaceRecognitionApp import FaceRecognitionApp
from wid import SlidePanel
import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
import mysql.connector
import os

class logedin:
    def __init__(self,parent,id=None,sub=None):
        def handle_click(event):
            if self.tree.identify_region(event.x, event.y) == "separator":
               return "break"
        def disp():
            def handle_click(self,event):
                if self.tree.identify_region(event.x, event.y) == "separator":
                   return "break"
            dis=tk.Toplevel(self.loged)
            dis.geometry("800x300")
            sql="SELECT login_id,Present,Arrived from face_id_data3"
            self.mycursor.execute(sql)
            res = self.mycursor.fetchall()
    
            frame = tk.Frame(dis)
            frame.pack(expand=True, fill='both')
    
            tree = ttk.Treeview(frame, columns=(1, 2, 3), show="headings", height=len(res))
            tree.pack(side='left', fill='both', expand=True)
            tree.heading(1, text="Login ID")
            tree.heading(2, text="Present")
            tree.heading(3, text="Arrived time")
    
            for row in res:
                tree.insert('', 'end', values=row)
            
            self.tree.bind('<Button-1>', handle_click)
            dis.mainloop()
                
        def download():
            def conf():
                # Query to select the Excel data from the database
                select_query = "SELECT excel_data FROM facedata4 where admin_id=%s and subject=%s"
                val=(self.entered_id,clicked.get())
                self.mycursor.execute(select_query,val)
                data = self.mycursor.fetchone()
                if data:
                    # Create a new Excel workbook
                    wb = Workbook()
                    ws = wb.active
                    
                    # Write the retrieved data to the Excel file
                    ws.title = "Face Data"
                    ws.append(["login_id", "Present", "Arrived", "Lecture"])  # Header row
                    wb_data = data[0]  # Extract binary data
                    with open('retrieved_face_data.xlsx', 'wb') as f:
                        f.write(wb_data)
                    messagebox.showinfo("Downloaded","Successful",parent=self.loged)
                else:
                    messagebox.showerror("Error", "Please enter valid ID and password",parent=self.loged)
            sub=tk.Toplevel(self.loged)
            sql="SELECT subject from facedata4 WHERE admin_id=%s"
            val=(self.entered_id,)
            self.mycursor.execute(sql,val)
            options = self.mycursor.fetchall()
            clicked = StringVar()
            clicked.set("")
            style = ttk.Style()
            style.configure("TMenubutton", background="#01589e")
            style.configure("TCombobox", background="#01589e", foreground='red')
            drop = ttk.OptionMenu(sub, clicked, *options).pack()
            #drop.place(relx=0.69,rely=0.576)
            #drop.config(width=30)
            co=tk.Button(sub,text="confirm",command=conf).pack()
            sub.mainloop()
        def emai():
            pass
        self.opne=None
        self.mydb = mysql.connector.connect(host="localhost", user="root", password="", database="facee")
        self.mycursor = self.mydb.cursor()
        self.parent=parent
        self.loged=parent
        self.loged.title("Loged")
        self.loged.configure(bg='#A8DEEA')
        self.loged.attributes('-fullscreen', True)
        today = datetime.date.today()
        day_of_week = today.weekday()
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        self.today_day = days[day_of_week]
        self.current_time = datetime.datetime.now().time()


        sql="SELECT admin_id,time,day,subject,branch,sem from time_tabel WHERE admin_id=%s and day=%s"
        val=('101','Friday')
        self.mycursor.execute(sql,val)
        res = self.mycursor.fetchall()

        query = "SELECT time, duratin FROM time_tabel where day=%s"
        val1=(self.today_day,)
        self.mycursor.execute(query,val1)
        time_data = self.mycursor.fetchall()
        
        self.time_list = []
        
        for time_start, duration in time_data:
            time_start = datetime.datetime.strptime(time_start, '%H:%M:%S').time()
            hour_int = time_start.hour
            minutes_int = time_start.minute
            self.time_list.append((hour_int, minutes_int, duration))

        image_path = os.path.abspath("bg4.jpg")
        bgimage_original = Image.open(image_path)
        bgimage_resized = bgimage_original.resize((self.loged.winfo_screenwidth(), self.loged.winfo_screenheight()))
        self.bg_photo_loged = ImageTk.PhotoImage(bgimage_resized)
    
        bglabel = tk.Label(self.loged, image=self.bg_photo_loged)
        bglabel.place(x=0, y=0, relwidth=1, relheight=1)
        bglabel.bind('<Configure>', self.resize_image)

        frame = tk.Frame(self.loged)
        frame.place(relx=0.1, rely=0.3, relwidth=0.8, relheight=0.5)

        ani=SlidePanel(self.loged,0, -0.3)
        
        bg2=ctk.CTkLabel(master=ani,text="",bg_color='#276196').pack(expand='True',fill = 'both')
        bu=ctk.CTkButton(master=ani,text='Download Attendance',bg_color='#276196',font=('Helvetica bold', 13),command=download).place(relx=0.05,rely=0.05,relwidth=0.9,relheight=0.12)
        bu1=ctk.CTkButton(master=ani,text='Send Attendance to E-mail',bg_color='#276196',font=('Helvetica bold', 13)).place(relx=0.05,rely=0.25,relwidth=0.9,relheight=0.12)
        bu2=ctk.CTkButton(master=ani,text='Change Schedule',bg_color='#276196',font=('Helvetica bold', 13)).place(relx=0.05,rely=0.45,relwidth=0.9,relheight=0.12)
        bu3=ctk.CTkButton(master=ani,text='Display',bg_color='#276196',font=('Helvetica bold', 13),command=disp).place(relx=0.05,rely=0.65,relwidth=0.9,relheight=0.12)


        self.tree = ttk.Treeview(frame, columns=(1, 2, 3, 4, 5, 6), show="headings", height=len(res))
        self.tree.grid(row=0, column=0, sticky="nsew")  # Use grid() instead of pack()

        self.tree.heading(1, text="Admin ID")
        self.tree.heading(2, text="Time")
        self.tree.heading(3, text="Day")
        self.tree.heading(4, text="Subject")
        self.tree.heading(5, text="Branch")
        self.tree.heading(6, text="Sem")
  
        for row in res:
            self.tree.insert('', 'end', values=row)
        
        self.tree.bind('<Button-1>', handle_click)
        self.tree.bind("<ButtonRelease-1>", self.handle_cell_click)
        
        scrollbar = ttk.Scrollbar(frame, orient='vertical', command=self.tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        scrollbar1 = ttk.Scrollbar(frame, orient='horizontal', command=self.tree.xview)
        scrollbar1.grid(row=1, column=0, sticky="ew") 
        self.tree.configure(xscrollcommand=scrollbar1.set)
        
        # Configure row and column weights to allow expansion
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        #self.dowm = tk.Button(self.loged, text="Download", command=down)
        #self.dowm.pack()

        self.loged.bind('<Escape>', self.close)

        self.backs = tk.Button(self.loged, text="Admin Page",bg='#0a94b0', command=self.back_to_login)
        self.backs.place(relx=0.01, rely=0.03, relwidth=0.1,relheight=0.07)

        self.ani=tk.Button(self.loged, text="=",bg='#0a94b0',command=ani.animate)
        self.ani.place(relx=0.12, rely=0.03, relwidth=0.04,relheight=0.07)

        self.mydb.commit()
    def close(self,event):
        try:
            self.loged.destroy()
        except:
            pass    
    def resize_image(self, event):
        new_width = event.width
        new_height = event.height
        self.bg_image_resized = self.bg_image_original.resize((new_width, new_height))
        self.bg_photo = ImageTk.PhotoImage(self.bg_image_resized)
        self.bg_label.config(image=self.bg_photo)

    def handle_cell_click(self,event):
        def che():
            self.time_list.remove((start_hour, start_minute, duration))
            #self.new=tk.Toplevel(self.loged)
            self.opne = FaceRecognitionApp(self.loged,  callback=lambda: self.on_face_recognition_close(lec_info,id,bra,sem))
            print("Opening Face Recognition App...")
        def no():
            ta.destroy()
        region = event.widget.identify_region(event.x, event.y)
        if region == "cell":
            item = self.tree.identify_row(event.y)
            self.values = self.tree.item(item, 'values')
            id=self.values[0]
            clicked_time = self.values[1]
            lec_info=self.values[3]
            bra=self.values[4]
            sem=self.values[5]
            # Convert the clicked_time string to a tuple of integers
            clicked_hour, clicked_minute, _ = map(int, clicked_time.split(":"))
            clicked_time_tuple = (clicked_hour, clicked_minute)
            found = False
            for start_hour, start_minute, duration in self.time_list:
                if (start_hour, start_minute) == clicked_time_tuple:
                    start_time = datetime.datetime.combine(datetime.date.today(), datetime.time(start_hour, start_minute))
                    end_time = start_time + datetime.timedelta(hours=int(duration))
                    if start_time.time() <= self.current_time <= end_time.time():
                        ta = tk.Toplevel(self.loged)
                        law = tk.Label(ta,text="do you want to markattendance?")
                        law.pack()
                        ba = tk.Button(ta,text="yess",command=che)
                        ba.pack()
                        ba1 = tk.Button(ta,text="no",command=no)
                        ba1.pack()
                        found = True
                        break
                    if not found:
                        messagebox.showerror("Lecture","Not time for this lec")
    def handle_cell_click(self,event):
        def che():
            self.time_list.remove((start_hour, start_minute, duration))
            #self.new=tk.Toplevel(self.loged)
            self.opne = FaceRecognitionApp(self.loged,  callback=lambda: self.on_face_recognition_close(lec_info,id,bra,sem))
            print("Opening Face Recognition App...")
        def no():
            ta.destroy()
        region = event.widget.identify_region(event.x, event.y)
        if region == "cell":
            item = self.tree.identify_row(event.y)
            self.values = self.tree.item(item, 'values')
            id=self.values[0]
            clicked_time = self.values[1]
            lec_info=self.values[3]
            bra=self.values[4]
            sem=self.values[5]
            # Convert the clicked_time string to a tuple of integers
            clicked_hour, clicked_minute, _ = map(int, clicked_time.split(":"))
            clicked_time_tuple = (clicked_hour, clicked_minute)
            found = False
            for start_hour, start_minute, duration in self.time_list:
                if (start_hour, start_minute) == clicked_time_tuple:
                    start_time = datetime.datetime.combine(datetime.date.today(), datetime.time(start_hour, start_minute))
                    end_time = start_time + datetime.timedelta(hours=int(duration))
                    if start_time.time() <= self.current_time <= end_time.time():
                        ta = tk.Toplevel(self.loged)
                        law = tk.Label(ta,text="do you want to markattendance?")
                        law.pack()
                        ba = tk.Button(ta,text="yess",command=che)
                        ba.pack()
                        ba1 = tk.Button(ta,text="no",command=no)
                        ba1.pack()
                        found = True
                        break
                    if not found:
                        messagebox.showinfo("Lecture", "Not time for this lec", parent=self.loged)
    def on_face_recognition_close(self,lec,idd,bra,sem):
        query = "SELECT login_id, Present, Arrived FROM face_id_data3"
        self.mycursor.execute(query)
        data = self.mycursor.fetchall()
        current_datetime = datetime.datetime.now()
        
        q1uery = "SELECT excel_data FROM facedata4 WHERE admin_id=%s AND subject=%s AND branch=%s AND sem=%s"
        self.mycursor.execute(q1uery, (idd, lec, bra, sem))
        existing_record = self.mycursor.fetchone()
        
        lecture_info = lec
        
        df = pd.DataFrame(data, columns=['login_id', 'Present', 'Arrived'])
        
        df['Lecture'] = lecture_info
        
        if existing_record:
            excel_data = existing_record[0]
            with io.BytesIO(excel_data) as bio:
                wb = load_workbook(filename=bio)
                ws = wb.active
        
                start_row = ws.max_row + 1
        
                current_date = current_datetime.now().strftime("%Y-%m-%d")
                ws.insert_rows(start_row)
                ws.cell(row=start_row, column=1, value=f"Day Separator - {current_date}")
        
                start_row += 1
        
                data_to_append = df.values.tolist()
        
                for row in data_to_append:
                    ws.append(row)
        
                with io.BytesIO() as bio:
                    wb.save(bio)
                    excel_data = bio.getvalue()
        
                update_query = "UPDATE facedata4 SET excel_data=%s WHERE admin_id=%s AND subject=%s AND branch=%s AND sem=%s"
        
                self.mycursor.execute(update_query, (excel_data, idd, lec, bra, sem))
                self.mydb.commit()
                print("Excel data updated in the database table facedata4")
                self.mycursor.execute("UPDATE face_id_data3 SET Present = '0'")
                self.mydb.commit()
        else:
            wb = Workbook()
            ws = wb.active
        
            current_date = current_datetime.now().strftime("%Y-%m-%d")
            ws.insert_rows(0) 
            ws.cell(row=1, column=1, value=f"Day Separator - {current_date}")
        
            data_to_append = df.values.tolist()
        
            for row in data_to_append:
                ws.append(row)
        
            with io.BytesIO() as bio:
                wb.save(bio)
                excel_data = bio.getvalue()
        
            insert_query = "INSERT INTO facedata4 (admin_id, subject, branch, sem, excel_data) VALUES (%s, %s, %s, %s, %s)"
            self.mycursor.execute(insert_query, (idd, lec, bra, sem,excel_data))
            self.mydb.commit()
            print("Excel data inserted into the database table facedata4")   
            self.mycursor.execute("UPDATE face_id_data3 SET Present = '0'")
            self.mydb.commit()

    
    def back_to_login(self):
        self.parent.start()
        

if __name__ == "__main__":
    rr=tk.Tk()
    app = logedin(parent=rr,id='101')
    rr.mainloop()