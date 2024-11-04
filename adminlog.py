from tkcalendar import Calendar
import io
import tkinter as tk
from tkinter import StringVar, ttk
import customtkinter as ctk
import os
import mysql.connector
from PIL import Image, ImageTk, ImageFilter, ImageDraw
from css import *
from datetime import datetime
import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from FaceRecognitionApp import FaceRecognitionApp
from wid import SlidePanel
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from io import BytesIO
class adminlog:
    def __init__(self,parent) -> None:
        self.parent = parent
        self.mydb = mysql.connector.connect(host="localhost", user="root", password="", database="facee")
        self.mycursor = self.mydb.cursor()
        self.bg_photo = None
        self.opne=None
        self.logins()
    def logins(self):
        self.root = ctk.CTk()
        self.root.title("Face recognition")
        self.root.configure(bg='#A8DEEA')
        self.root.attributes('-fullscreen', True)  # Fullscreen mode
        
        image_path = os.path.abspath("Frame2.png")
        self.bg_image_original = Image.open(image_path)
        self.bg_image_resized = self.bg_image_original.resize((self.root.winfo_screenwidth(), self.root.winfo_screenheight()))
        self.bg_photo = ImageTk.PhotoImage(self.bg_image_resized)
    
        # Display the background image using a label
        self.bg_label = tk.Label(self.root, image=self.bg_photo)
        self.bg_label.place(x=0, y=0, relwidth=1, relheight=1)
        self.bg_label.bind('<Configure>', self.resize_image)
        self.namee =  tk.Entry(self.root, font=("Arial", 18))
        self.namee.place(relx=0.4453, rely=0.44,relwidth=0.257,relheight=0.06)
        
        self.passw = tk.Entry(self.root, font=("Arial", 18),show='*')
        self.passw.place(relx=0.4453, rely=0.546,relwidth=0.257,relheight=0.06)
        gradient_photo = self.resize_imageb('button2.png', relwidth=1, relheight=1)
        bu = get_button(self.root, None, color=None, image=gradient_photo, command=self.lin, fg="black")
        bu.place(relx=0.374996, rely=0.7247, relwidth=0.2359, relheight=0.078)

        relw=35
        relh=20
        gradient_photo1=self.resize_imageb('button1.jpg', relw, relh)

        self.btm =  tk.Button(self.root,text="Log out",image= gradient_photo1 , compound='center',command=self.back_to_main, bg='#0d6e81', activebackground='#6DA5C0')
        self.btm.place(relx=0.01, rely=0.03, relwidth=0.1,relheight=0.07)
        self.btm.config(image=gradient_photo1)
        
        self.ani=tk.Button(self.root, text="=",bg='#0a94b0',command=None)
        self.root.bind('<Escape>', self.close)
        self.root.mainloop()

    def back_to_main(self):
        self.root.destroy()
        self.parent.create_main_window()

    def lin(self):
        admins = {"101": '12333', "102": '1233', "103": '1233', "104": '1233'}
        try:
            self.entered_id = self.namee.get()  
            entered_password = self.passw.get()
            if (self.entered_id.strip() in admins.keys()) and admins[self.entered_id]==entered_password:
                messagebox.showinfo("Welcome", "Successfully logged in")
                self.logined()
            else:
                messagebox.showerror("Not Admin", "Please try with correct ID and password",parent=self.root)
        except ValueError:
            messagebox.showerror("Error", "Please enter valid ID and password",parent=self.root)

    def resize_imageb(self,image_path, relwidth, relheight):
        # Open the image file using PIL and preserve the original color space and gamma information
        original_image = Image.open(image_path).convert('RGB')
        
        # Calculate the new dimensions based on relative width and height
        new_width = int(original_image.width * relwidth)
        new_height = int(original_image.height * relheight)
        
        # Resize the image
        resized_image = original_image.resize((new_width, new_height), Image.LANCZOS)
        
        # Convert the resized image to a PhotoImage object using ImageTk.PhotoImage
        return ImageTk.PhotoImage(resized_image)

    def resize_image(self, event):
        new_width = event.width
        new_height = event.height
        self.bg_image_resized = self.bg_image_original.resize((new_width, new_height))
        self.bg_photo = ImageTk.PhotoImage(self.bg_image_resized)
        self.bg_label.config(image=self.bg_photo)
    
    def close(self,event):
        try:
            self.root.destroy()
        except:
            pass

    def logined(self):
        def handle_click(event):
            if self.tree.identify_region(event.x, event.y) == "separator":
               return "break"
        def refresh_loged():
            # Clear existing data
            for i in self.tree.get_children():
                self.tree.delete(i)
            
            # Re-query the database
            sql = "SELECT admin_id, time, day, subject, branch, sem from time_tabel WHERE admin_id=%s and day=%s"
            val = (self.entered_id, self.today_day)
            self.mycursor.execute(sql, val)
            res = self.mycursor.fetchall()
        
            # Re-populate the tree view
            for row in res:
                self.tree.insert('', 'end', values=row)
        def disp():
            def handle_click(self,event):
                if self.tree.identify_region(event.x, event.y) == "separator":
                   return "break"
            dis=tk.Toplevel(self.loged)
            dis.geometry("800x300")
            sql="SELECT login_id,Present,Arrived from face_id_data3"
            self.mycursor.execute(sql)
            res = self.mycursor.fetchall()
    
            frame = tk.Frame(dis)
            frame.pack(expand=True, fill='both')
    
            tree = ttk.Treeview(frame, columns=(1, 2, 3), show="headings", height=len(res))
            tree.pack(side='left', fill='both', expand=True)
            tree.heading(1, text="Login ID")
            tree.heading(2, text="Present")
            tree.heading(3, text="Arrived time")
    
            for row in res:
                tree.insert('', 'end', values=row)
            
            self.tree.bind('<Button-1>', handle_click)
            dis.mainloop()
                
        def download():
            def conf():
                # Query to select the Excel data from the database
                select_query = "SELECT excel_data FROM facedata4 where admin_id=%s and subject=%s"
                val=(self.entered_id,clicked.get())
                self.mycursor.execute(select_query,val)
                data = self.mycursor.fetchone()
                if data:
                    # Create a new Excel workbook
                    wb = Workbook()
                    ws = wb.active
                    
                    # Write the retrieved data to the Excel file
                    ws.title = "Face Data"
                    ws.append(["login_id", "Present", "Arrived", "Lecture"])  # Header row
                    wb_data = data[0]  # Extract binary data

                    with open(f'{clicked.get()}database.xlsx', 'wb') as f:
                        f.write(wb_data)
                    messagebox.showinfo("Downloaded","Successful",parent=self.loged)
                else:
                    messagebox.showerror("Error", "No data present",parent=self.loged)
            
            sql = "SELECT subject FROM facedata4 WHERE admin_id=%s"
            val = (self.entered_id,)
            self.mycursor.execute(sql, val)
            sub=tk.Toplevel(self.loged)
            sub.geometry('300x200')
            bg_image_original = Image.open("bgdown.jpg")
            bg_image_resized = bg_image_original.resize((300,200))
            bg_photo = ImageTk.PhotoImage(bg_image_resized)
            
            bg_label = tk.Label(sub, image=bg_photo)
            bg_label.place(relx=0, rely=0, relwidth=1, relheight=1)
            bg_label.bind('<Configure>', self.resize_image)
            bg_label.image = bg_photo
            
            options = [item[0] for item in self.mycursor.fetchall()]  
            # Dropdown for selecting subjects
            clicked = tk.StringVar(sub)
            clicked.set("Choose a subject")  # default value
            style = ttk.Style()
            style.configure("TMenubutton", background="#01589e")
            style.configure("TCombobox", background="#01589e", foreground='red')
            drop = ttk.OptionMenu(sub, clicked, clicked.get(), *options)
            drop.place(relx=0.17,rely=0.2,relheight=0.16,relwidth=0.7)
            co=tk.Button(sub,text="confirm",command=conf)
            co.place(relx=0.17,rely=0.6,relheight=0.2,relwidth=0.7)
            sub.mainloop()
                    
        def emai():
            def conf():
                # Query to select the Excel data from the database
                select_query = "SELECT excel_data FROM facedata4 WHERE admin_id=%s AND subject=%s"
                val = (self.entered_id, clicked.get())
                self.mycursor.execute(select_query, val)
                data = self.mycursor.fetchone()
                if data:
                    # Create a new Excel workbook in memory
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Face Data"
                    ws.append(["login_id", "Present", "Arrived", "Lecture"])  # Header row
                    wb_data = data[0]  # Extract binary data
                    bio = BytesIO(wb_data)  # Use BytesIO object to handle the Excel data in memory
        
                    # Send the Excel file via email
                    receiver_email = to.get()
                    send_email(bio, app_password="nbri nljt jiqk adsy", receiver_email=receiver_email)
                else:
                    messagebox.showerror("Error", "Data not found for provided ID and subject", parent=sub)
        
            def send_email(file_stream, receiver_email, app_password, sender_email='soham56kadam@gmail.com'):
                message = MIMEMultipart()
                message['From'] = sender_email
                message['To'] = receiver_email
                message['Subject'] = 'SUCCSSFULL'
        
                part = MIMEBase('application', "octet-stream")
                part.set_payload(file_stream.getvalue())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', 'attachment; filename="FaceData.xlsx"')
                message.attach(part)
        
                server = smtplib.SMTP('smtp.gmail.com', 587)
                server.starttls()
                server.login(sender_email, app_password)
                server.sendmail(sender_email, receiver_email, message.as_string())
                server.quit()
        
                messagebox.showinfo("Email Sent", "Excel file has been sent successfully!", parent=sub)
        
            # Query to fetch subjects
            sql = "SELECT subject FROM facedata4 WHERE admin_id=%s"
            val = (self.entered_id,)
            self.mycursor.execute(sql, val)
            options = [item[0] for item in self.mycursor.fetchall()]
        
            # Create the main window
            sub = tk.Toplevel(self.loged)
            sub.geometry('550x400')
        
            # Load the background image
            bg_image_original = Image.open("bgche.png")
            bg_image_resized = bg_image_original.resize((550, 400))
            bg_photo = ImageTk.PhotoImage(bg_image_resized)
        
            # Display the background image using a label
            bg_label = tk.Label(sub, image=bg_photo)
            bg_label.place(relx=0, rely=0, relwidth=1, relheight=1)
            bg_label.bind('<Configure>', self.resize_image)
            # Save a reference to the photo to avoid garbage collection
            bg_label.image = bg_photo
        
            to = ttk.Entry(sub, font=('Helvetica bold', 14))
            to.insert(0, "Enter email address to send")
            to.place(relx=0.36, rely=0.27, relwidth=0.6, relheight=0.1)
        
            # Dropdown for selecting subjects
            clicked = tk.StringVar(sub)
            clicked.set("Choose a subject")  # default value
        
            drop = ttk.OptionMenu(sub, clicked, clicked.get(), *options)
            drop.place(relx=0.36, rely=0.44, relheight=0.078, relwidth=0.6)
        
            # Confirmation button
            co = tk.Button(sub, text="confirm", fg='white', font=('Times Roman', 13), bg='#0000ff', command=conf)
            co.place(relx=0.28, rely=0.66, relheight=0.1, relwidth=0.45)
        
            # Main application loop
            sub.mainloop()        
        def ChangeSch():
            def popup_calendar(id):
                top = tk.Toplevel(sub)
                cal = Calendar(top, selectmode='day')
                cal.pack(pady=20, padx=20)
            
                def on_date_select():
                    selected_date = cal.selection_get()
                    day_of_week = selected_date.strftime('%A')
                    formatted_date = selected_date.strftime('%Y-%m-%d')
                    date_entry.delete(0, tk.END)
                    date_entry.insert(0,formatted_date)
                    update_options(selected_date,id)
                    top.destroy()
            
                ok_btn = tk.Button(top, text="OK", command=on_date_select)
                ok_btn.pack(pady=10)
            
            def compare_and_update_tables():
                    query = """
                    SELECT t.time, t.day, t.admin_id AS admin_id_time_table, r.admin_id AS admin_id_rollbk
                    FROM time_tabel t
                    INNER JOIN rollbk r ON t.time = r.time AND t.day = r.day
                    WHERE t.admin_id <> r.admin_id;
                    """
                    self.mycursor.execute(query)
                    discrepancies = self.mycursor.fetchall()
                
                    if discrepancies:
                        updates_made = 0
                        for discrepancy in discrepancies:
                            # Access elements by index
                            admin_id_rollbk = discrepancy[3]
                            time = discrepancy[0]
                            day = discrepancy[1]
                
                            # Update time_tabel to use admin_id from rollbk
                            update_query = """
                            UPDATE time_tabel
                            SET admin_id = %s
                            WHERE time = %s AND day = %s;
                            """
                            self.mycursor.execute(update_query, (admin_id_rollbk, time, day))
                            updates_made += self.mycursor.rowcount
                
                        self.mydb.commit()
                        messagebox.showinfo("Comparison Result", f"Updates made {updates_made}.", parent=sub)
                        refresh_loged() 
                    else:
                        messagebox.showinfo("Comparison Result", "No changes found.", parent=sub)
                        refresh_loged() 
                                
            def update_options(selected_date,id):
                day_of_week = selected_date.strftime('%A')
                sql = "SELECT time FROM time_tabel WHERE admin_id=%s AND day=%s"
                val = (id, day_of_week)
                self.mycursor.execute(sql, val)
                new_options = [item[0] for item in self.mycursor.fetchall()]
                menu = drop['menu']
                menu.delete(0, 'end')
                for option in new_options:
                    menu.add_command(label=option, command=lambda value=option: clicked.set(value))
                if new_options:
                    clicked.set(new_options[0])
                else:
                    clicked.set('No times available')
            
            def conf():
                update="UPDATE time_tabel SET admin_id=%s WHERE time=%s"
                val=(to.get(),clicked.get())
                self.mycursor.execute(update,val)
                self.mydb.commit()
                sub.destroy()
                messagebox.showinfo("Successful","Schedule change is done",parent=sub)
                refresh_loged()

            sub = tk.Toplevel(self.loged)
            sub.geometry('550x400')
            bg_image_original = Image.open("bg4change.jpg")
            bg_image_resized = bg_image_original.resize((550, 400))
            bg_photo = ImageTk.PhotoImage(bg_image_resized)
            
            bg_label = tk.Label(sub, image=bg_photo)
            bg_label.place(relx=0, rely=0, relwidth=1, relheight=1)
            bg_label.bind('<Configure>', self.resize_image)
            bg_label.image = bg_photo
            
            to=tk.Entry(sub,font=('Helvetica bold', 12))
            to.place(relx=0.305,rely=0.3,relwidth=0.5,relheight=0.083)
            
            date_entry = tk.Entry(sub, font=('Helvetica bold', 12))
            date_entry.place(relx=0.305, rely=0.42, relwidth=0.5, relheight=0.07)
            date_entry.bind("<Button-1>", lambda event: popup_calendar(id=self.entered_id))
            
            clicked = tk.StringVar(sub)
            clicked.set("Choose a time")
            
            drop = ttk.OptionMenu(sub, clicked, "Choose a time", ())
            drop.place(relx=0.306, rely=0.526, relheight=0.078, relwidth=0.4)
            
            co = tk.Button(sub, text="Confirm", bg='#0a94b0', command=conf)
            co.place(relx=0.5, rely=0.7, relheight=0.1, relwidth=0.35)
            
            re = tk.Button(sub, text="Reset time table", bg='#0a94b0', command=compare_and_update_tables)
            re.place(relx=0.13, rely=0.7, relheight=0.1, relwidth=0.35)
            
            sub.mainloop()
        
        self.loged = tk.Toplevel(self.root)
        self.loged.title("Loged")
        self.loged.configure(bg='#A8DEEA')
        self.loged.attributes('-fullscreen', True)
        today = datetime.date.today()
        day_of_week = today.weekday()
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        self.today_day = days[day_of_week]
        self.current_time = datetime.datetime.now().time()
        sql="SELECT admin_id,time,day,subject,branch,sem from time_tabel WHERE admin_id=%s and day=%s"
        val=(self.entered_id,self.today_day)
        self.mycursor.execute(sql,val)
        res = self.mycursor.fetchall()

        query = "SELECT time, duratin FROM time_tabel where day=%s"
        val1=(self.today_day,)
        self.mycursor.execute(query,val1)
        time_data = self.mycursor.fetchall()
        
        self.time_list = []
        
        for time_start, duration in time_data:
            time_start = datetime.datetime.strptime(time_start, '%H:%M:%S').time()
            hour_int = time_start.hour
            minutes_int = time_start.minute
            self.time_list.append((hour_int, minutes_int, duration))

        image_path = os.path.abspath("bg4.jpg")
        bgimage_original = Image.open(image_path)
        bgimage_resized = bgimage_original.resize((self.loged.winfo_screenwidth(), self.loged.winfo_screenheight()))
        self.bg_photo_loged = ImageTk.PhotoImage(bgimage_resized)
    
        bglabel = tk.Label(self.loged, image=self.bg_photo_loged)
        bglabel.place(x=0, y=0, relwidth=1, relheight=1)
        bglabel.bind('<Configure>', self.resize_image)

        frame = tk.Frame(self.loged)
        frame.place(relx=0.1, rely=0.3, relwidth=0.8, relheight=0.5)

        ani=SlidePanel(self.loged,0, -0.3)
        
        bg2=ctk.CTkLabel(master=ani,text="",bg_color='#276196').pack(expand='True',fill = 'both')
        bu=ctk.CTkButton(master=ani,text='Download Attendance',bg_color='#276196',font=('Helvetica bold', 16),command=download).place(relx=0.05,rely=0.05,relwidth=0.9,relheight=0.12)
        bu1=ctk.CTkButton(master=ani,text='Send Attendance to E-mail',bg_color='#276196',font=('Helvetica bold', 16),command=emai).place(relx=0.05,rely=0.25,relwidth=0.9,relheight=0.12)
        bu2=ctk.CTkButton(master=ani,text='Change Schedule',bg_color='#276196',font=('Helvetica bold', 16),command=ChangeSch).place(relx=0.05,rely=0.45,relwidth=0.9,relheight=0.12)
        bu3=ctk.CTkButton(master=ani,text='Display',bg_color='#276196',font=('Helvetica bold', 16),command=disp).place(relx=0.05,rely=0.65,relwidth=0.9,relheight=0.12)
                
        self.tree = ttk.Treeview(frame, columns=(1, 2, 3, 4, 5, 6), show="headings", height=len(res))
        self.tree.grid(row=0, column=0, sticky="nsew")

        self.tree.heading(1, text="Admin ID")
        self.tree.heading(2, text="Time")
        self.tree.heading(3, text="Day")
        self.tree.heading(4, text="Subject")
        self.tree.heading(5, text="Branch")
        self.tree.heading(6, text="Sem")
  
        for row in res:
            self.tree.insert('', 'end', values=row)
        
        self.tree.bind('<Button-1>', handle_click)
        self.tree.bind("<ButtonRelease-1>", self.handle_cell_click)
        
        scrollbar = ttk.Scrollbar(frame, orient='vertical', command=self.tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        scrollbar1 = ttk.Scrollbar(frame, orient='horizontal', command=self.tree.xview)
        scrollbar1.grid(row=1, column=0, sticky="ew") 
        self.tree.configure(xscrollcommand=scrollbar1.set)
        
        # Configure row and column weights to allow expansion
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        self.backs = tk.Button(self.loged, text="Admin Page",bg='#0a94b0', command=self.back_to_login)
        self.backs.place(relx=0.01, rely=0.03, relwidth=0.1,relheight=0.07)

        self.ani=tk.Button(self.loged, text="=",bg='#0a94b0',command=ani.animate)
        self.ani.place(relx=0.12, rely=0.03, relwidth=0.04,relheight=0.07)

        self.mydb.commit()

    def handle_cell_click(self,event):
        def che():
            ta.destroy()
            self.time_list.remove((start_hour, start_minute, duration))
            self.opne = FaceRecognitionApp(self.loged,callback=lambda: self.on_face_recognition_close(lec_info,id,bra,sem),bra=bra,sem=sem)
        def no():
            ta.destroy()
        region = event.widget.identify_region(event.x, event.y)
        if region == "cell":
            item = self.tree.identify_row(event.y)
            self.values = self.tree.item(item, 'values')
            id=self.values[0]
            clicked_time = self.values[1]
            lec_info=self.values[3]
            bra=self.values[4]
            sem=self.values[5]
            # Convert the clicked_time string to a tuple of integers
            clicked_hour, clicked_minute, _ = map(int, clicked_time.split(":"))
            clicked_time_tuple = (clicked_hour, clicked_minute)
            found = False
            for start_hour, start_minute, duration in self.time_list:
                if (start_hour, start_minute) == clicked_time_tuple:
                    start_time = datetime.datetime.combine(datetime.date.today(), datetime.time(start_hour, start_minute))
                    end_time = start_time + datetime.timedelta(hours=int(duration))
                    if start_time.time() <= self.current_time <= end_time.time():
                        ta = tk.Toplevel(self.loged)
                        law = tk.Label(ta,text="Do you want to markattendance?")
                        law.pack()
                        ba = tk.Button(ta,text="yess",command=che)
                        ba.pack()
                        ba1 = tk.Button(ta,text="no",command=no)
                        ba1.pack()
                        found = True
                        break
            if not found:
                messagebox.showerror("Lecture","Not time for this lec",parent=self.loged)
    
    def on_face_recognition_close(self,lec,idd,bra,sem):
        query = "SELECT login_id, Present, Arrived FROM face_id_data3 WHERE Department=%s and sem=%s"
        val=(bra,sem)
        self.mycursor.execute(query,val)
        data = self.mycursor.fetchall()
        current_datetime = datetime.datetime.now()
        
        q1uery = "SELECT excel_data FROM facedata4 WHERE admin_id=%s AND subject=%s AND branch=%s AND sem=%s"
        self.mycursor.execute(q1uery, (idd, lec, bra, sem))
        existing_record = self.mycursor.fetchone()
        
        lecture_info = lec
        
        df = pd.DataFrame(data, columns=['login_id', 'Present', 'Arrived'])
        
        df['Lecture'] = lecture_info

        if existing_record and existing_record[0]:
            excel_data = existing_record[0]
            with io.BytesIO(excel_data) as bio:
                wb = load_workbook(filename=bio)
                ws = wb.active
        
                start_row = ws.max_row + 1
        
                current_date = current_datetime.now().strftime("%Y-%m-%d")
                ws.insert_rows(start_row)
                ws.cell(row=start_row, column=1, value=f"Day Separator - {current_date}")
        
                start_row += 1
        
                data_to_append = df.values.tolist()
        
                for row in data_to_append:
                    ws.append(row)
        
                with io.BytesIO() as bio:
                    wb.save(bio)
                    excel_data = bio.getvalue()
        
                update_query = "UPDATE facedata4 SET excel_data=%s WHERE admin_id=%s AND subject=%s AND branch=%s AND sem=%s"
        
                self.mycursor.execute(update_query, (excel_data, idd, lec, bra, sem))
                self.mydb.commit()
                self.mycursor.execute("UPDATE face_id_data3 SET Present = '0'")
                self.mydb.commit()
        else:
            wb = Workbook()
            ws = wb.active
        
            current_date = current_datetime.now().strftime("%Y-%m-%d")
            ws.insert_rows(0) 
            ws.cell(row=1, column=1, value=f"Day Separator - {current_date}")
        
            data_to_append = df.values.tolist()
        
            for row in data_to_append:
                ws.append(row)
        
            with io.BytesIO() as bio:
                wb.save(bio)
                excel_data = bio.getvalue()
        
            insert_query =  "UPDATE facedata4 SET excel_data=%s WHERE admin_id=%s AND subject=%s AND branch=%s AND sem=%s"
            self.mycursor.execute(insert_query, (excel_data,idd, lec, bra, sem))
            self.mydb.commit()
            self.mycursor.execute("UPDATE face_id_data3 SET Present = '0'")
            self.mydb.commit()
        self.logined()

    def back_to_login(self):
        self.loged.withdraw()
        self.root.deiconify()
    
    def start(self):
        self.root.mainloop()
    
if __name__ == "__main__":
    app = adminlog(parent=None)
    app.logined()